import pyshark
from datetime import datetime
import pandas as pd
from tqdm import tqdm
import math
import os
from urllib.parse import urlparse
from openpyxl import load_workbook

def categorize_url(url):
    return (
        "нет" if not url else
        "студент" if "mtuci" in (urlparse(url).netloc.lower() if "://" in url else url.lower()) else
        "видеохостинг" if any(x in (urlparse(url).netloc.lower() if "://" in url else url.lower()) for x in ["youtube", "vid", "video"]) else
        "информационный сайт" if any(x in (urlparse(url).netloc.lower() if "://" in url else url.lower()) for x in ["vk.com", "vk"]) else
        "нет"
    )

def expand_categories(df):
    rows = []
    for _, row in df.iterrows():
        categories = row['http_url_categories']
        if not categories or categories == "нет":
            rows.append(row)
        else:
            for cat in categories.split(';'):
                new_row = row.copy()
                new_row['http_url_categories'] = cat
                rows.append(new_row)
    return pd.DataFrame(rows)

def append_to_csv(df, output_file):
    header = not os.path.exists(output_file)
    df_expanded = expand_categories(df)
    df_expanded.to_csv(output_file, mode='a', index=False, header=header)

def append_to_excel(df, output_file, sheet_name="Sheet1"):
    df_expanded = expand_categories(df)
    if not os.path.exists(output_file):
        df_expanded.to_excel(output_file, index=False, sheet_name=sheet_name)
    else:
        book = load_workbook(output_file)
        with pd.ExcelWriter(output_file, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            startrow = book[sheet_name].max_row if sheet_name in book.sheetnames else 0
            df_expanded.to_excel(writer, index=False, sheet_name=sheet_name, startrow=startrow, header=False)

def save_data(df, csv_file=None, excel_file=None):
    if csv_file:
        append_to_csv(df, csv_file)
    if excel_file:
        append_to_excel(df, excel_file)

def save_and_reset(current_window_start, tcp_packets, udp_packets, http_packets,
                   tcp_sessions, udp_sessions, tcp_src_ports, udp_src_ports,
                   tcp_dst_ports, udp_dst_ports, all_ips, tcp_lengths, udp_lengths,
                   all_lengths, http_urls, http_url_cats,
                   time_windows, tcp_packet_counts, udp_packet_counts, http_packet_counts,
                   tcp_sessions_counts, udp_sessions_counts, tcp_src_ports_counts, udp_src_ports_counts,
                   tcp_dst_ports_counts, udp_dst_ports_counts, avg_pkt_lengths, avg_tcp_pkt_lengths,
                   avg_udp_pkt_lengths, unique_ip_counts, http_url_counts, avg_http_url_lengths, http_url_categories):
    
    time_label = datetime.fromtimestamp(current_window_start).strftime('%Y-%m-%d %H:%M:%S')

    time_windows.append(time_label)
    tcp_packet_counts.append(tcp_packets)
    udp_packet_counts.append(udp_packets)
    http_packet_counts.append(http_packets)
    tcp_sessions_counts.append(len(tcp_sessions))
    udp_sessions_counts.append(len(udp_sessions))
    tcp_src_ports_counts.append(len(tcp_src_ports))
    udp_src_ports_counts.append(len(udp_src_ports))
    tcp_dst_ports_counts.append(len(tcp_dst_ports))
    udp_dst_ports_counts.append(len(udp_dst_ports))
    avg_pkt_lengths.append(int(sum(all_lengths)/len(all_lengths)) if all_lengths else 0)
    avg_tcp_pkt_lengths.append(int(sum(tcp_lengths)/len(tcp_lengths)) if tcp_lengths else 0)
    avg_udp_pkt_lengths.append(int(sum(udp_lengths)/len(udp_lengths)) if udp_lengths else 0)
    unique_ip_counts.append(len(all_ips))
    http_url_counts.append(len(http_urls))
    avg_http_url_lengths.append(int(sum(len(url) for url in http_urls)/len(http_urls)) if http_urls else 0)
    http_url_categories.append(";".join(sorted(http_url_cats)) if http_url_cats else "нет")

    # Сброс данных
    tcp_packets = udp_packets = http_packets = 0
    tcp_sessions.clear()
    udp_sessions.clear()
    tcp_src_ports.clear()
    udp_src_ports.clear()
    tcp_dst_ports.clear()
    udp_dst_ports.clear()
    all_ips.clear()
    tcp_lengths.clear()
    udp_lengths.clear()
    all_lengths.clear()
    http_urls.clear()
    http_url_cats.clear()

    return (tcp_packets, udp_packets, http_packets,
            tcp_sessions, udp_sessions, tcp_src_ports, udp_src_ports,
            tcp_dst_ports, udp_dst_ports, all_ips, tcp_lengths, udp_lengths,
            all_lengths, http_urls, http_url_cats)

def process_pcap(pcap_file, output_csv=None, output_excel=None, save_interval=1000):
    cap = pyshark.FileCapture(pcap_file, keep_packets=False)

    current_window_start = None
    processed_packets = 0

    (time_windows, tcp_packet_counts, udp_packet_counts, http_packet_counts,
    tcp_sessions_counts, udp_sessions_counts, tcp_src_ports_counts, udp_src_ports_counts,
    tcp_dst_ports_counts, udp_dst_ports_counts, avg_pkt_lengths, avg_tcp_pkt_lengths,
    avg_udp_pkt_lengths, unique_ip_counts, http_url_counts, avg_http_url_lengths,
    http_url_categories) = ([] for _ in range(17))

    tcp_packets = udp_packets = http_packets = 0 
    tcp_sessions = set(); udp_sessions = set()
    tcp_src_ports = set(); udp_src_ports = set()
    tcp_dst_ports = set(); udp_dst_ports = set()
    all_ips = set(); 
    tcp_lengths = []; udp_lengths = []
    all_lengths = []
    http_urls = []; http_url_cats = set()

    for pkt in tqdm(cap, desc="Обработка пакетов", unit="pkt"):
        pkt_time = float(pkt.sniff_timestamp)
        pkt_sec = math.floor(pkt_time)

        if current_window_start is None:
            current_window_start = pkt_sec

        if pkt_sec != current_window_start:
            (tcp_packets, udp_packets, http_packets,
             tcp_sessions, udp_sessions, tcp_src_ports, udp_src_ports,
             tcp_dst_ports, udp_dst_ports, all_ips, tcp_lengths, udp_lengths,
             all_lengths, http_urls, http_url_cats) = save_and_reset(
                current_window_start, tcp_packets, udp_packets, http_packets,
                tcp_sessions, udp_sessions, tcp_src_ports, udp_src_ports,
                tcp_dst_ports, udp_dst_ports, all_ips, tcp_lengths, udp_lengths,
                all_lengths, http_urls, http_url_cats,
                time_windows, tcp_packet_counts, udp_packet_counts, http_packet_counts,
                tcp_sessions_counts, udp_sessions_counts, tcp_src_ports_counts, udp_src_ports_counts,
                tcp_dst_ports_counts, udp_dst_ports_counts, avg_pkt_lengths, avg_tcp_pkt_lengths,
                avg_udp_pkt_lengths, unique_ip_counts, http_url_counts, avg_http_url_lengths, http_url_categories
            )
            current_window_start = pkt_sec

        all_lengths.append(int(pkt.length))

        if 'IP' in pkt:
            all_ips.add(pkt.ip.src)
            all_ips.add(pkt.ip.dst)

        if 'TCP' in pkt:
            tcp_packets += 1
            tcp_lengths.append(int(pkt.length))
            src_ip = pkt.ip.src if 'IP' in pkt else None
            dst_ip = pkt.ip.dst if 'IP' in pkt else None
            src_port = int(pkt.tcp.srcport)
            dst_port = int(pkt.tcp.dstport)

            if src_ip and dst_ip:
                tcp_sessions.add((src_ip, src_port, dst_ip, dst_port))
            tcp_src_ports.add(src_port)
            tcp_dst_ports.add(dst_port)

            if 'HTTP' in pkt:
                http_packets += 1

        elif 'UDP' in pkt:
            udp_packets += 1
            udp_lengths.append(int(pkt.length))
            src_ip = pkt.ip.src if 'IP' in pkt else None
            dst_ip = pkt.ip.dst if 'IP' in pkt else None
            src_port = int(pkt.udp.srcport)
            dst_port = int(pkt.udp.dstport)

            if src_ip and dst_ip:
                udp_sessions.add((src_ip, src_port, dst_ip, dst_port))
            udp_src_ports.add(src_port)
            udp_dst_ports.add(dst_port)

        if 'DNS' in pkt:
            try:
                domain_name = pkt.dns.qry_name
                if domain_name:
                    http_urls.append(domain_name)
                    http_url_cats.add(categorize_url(domain_name))
            except AttributeError:
                pass

        processed_packets += 1

        if processed_packets % save_interval == 0 and time_windows:
            df = pd.DataFrame({
                'time_window_start': time_windows,
                'tcp_packet_count': tcp_packet_counts,
                'udp_packet_count': udp_packet_counts,
                'http_packet_count': http_packet_counts,
                'tcp_sessions_count': tcp_sessions_counts,
                'udp_sessions_count': udp_sessions_counts,
                'tcp_src_ports_count': tcp_src_ports_counts,
                'udp_src_ports_count': udp_src_ports_counts,
                'tcp_dst_ports_count': tcp_dst_ports_counts,
                'udp_dst_ports_count': udp_dst_ports_counts,
                'avg_pkt_length': avg_pkt_lengths,
                'avg_tcp_pkt_length': avg_tcp_pkt_lengths,
                'avg_udp_pkt_length': avg_udp_pkt_lengths,
                'unique_ip_count': unique_ip_counts,
                'http_url_count': http_url_counts,
                'avg_http_url_length': avg_http_url_lengths,
                'http_url_categories': http_url_categories
            })

            save_data(df, csv_file=output_csv, excel_file=output_excel)

            # Очистка после сохранения
            time_windows.clear()
            tcp_packet_counts.clear()
            udp_packet_counts.clear()
            http_packet_counts.clear()
            tcp_sessions_counts.clear()
            udp_sessions_counts.clear()
            tcp_src_ports_counts.clear()
            udp_src_ports_counts.clear()
            tcp_dst_ports_counts.clear()
            udp_dst_ports_counts.clear()
            avg_pkt_lengths.clear()
            avg_tcp_pkt_lengths.clear()
            avg_udp_pkt_lengths.clear()
            unique_ip_counts.clear()
            http_url_counts.clear()
            avg_http_url_lengths.clear()
            http_url_categories.clear()

    # Сохраняем остаток после обработки всех пакетов
    if time_windows:
        df = pd.DataFrame({
            'time_window_start': time_windows,
            'tcp_packet_count': tcp_packet_counts,
            'udp_packet_count': udp_packet_counts,
            'http_packet_count': http_packet_counts,
            'tcp_sessions_count': tcp_sessions_counts,
            'udp_sessions_count': udp_sessions_counts,
            'tcp_src_ports_count': tcp_src_ports_counts,
            'udp_src_ports_count': udp_src_ports_counts,
            'tcp_dst_ports_count': tcp_dst_ports_counts,
            'udp_dst_ports_count': udp_dst_ports_counts,
            'avg_pkt_length': avg_pkt_lengths,
            'avg_tcp_pkt_length': avg_tcp_pkt_lengths,
            'avg_udp_pkt_length': avg_udp_pkt_lengths,
            'unique_ip_count': unique_ip_counts,
            'http_url_count': http_url_counts,
            'avg_http_url_length': avg_http_url_lengths,
            'http_url_categories': http_url_categories
        })
        save_data(df, csv_file=output_csv, excel_file=output_excel)

    print(f"Данные сохранены в {output_csv if output_csv else ''} {output_excel if output_excel else ''}")

if __name__ == "__main__":
    pcap_file = '../yu_dump.pcapng'
    output_csv = '../datasets/yu_dataset.csv'   # или None, если не нужен CSV
    output_excel = '../datasets/yu_dataset.xlsx'  # или None, если не нужен Excel
    save_interval = 1000

    process_pcap(pcap_file, output_csv=output_csv, output_excel=output_excel, save_interval=save_interval)
